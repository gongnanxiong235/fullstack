# author:gongnanxiong
# date:2018/11/12

import openpyxl
from openpyxl import Workbook
def writeExcel(path):
    wb=Workbook()
    sheet=wb.active
    sheet.title='test Excel2'
    value = [["名称", "价格", "出版社", "语言"],
                 ["如何高效读懂一本书", "22.93", "机械工业出版社", "中文"],
                 ["暗时间", "32.4", "人民邮电出版社", "中文"],
                 ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for row in range(len(value)):
        for column in range(len(value[row])):
            sheet.cell(row+1,column+1,value[row][column])
    wb.save(path)
    print('写入成功')
#writeExcel('2007.xlsx

def readExcel(path):
    wx=openpyxl.load_workbook(path) #type:Workbook
    sheet=wx["test Excel2"]
    print(sheet)
    cell_list = []
    for row in sheet.rows:
        row_list = []
        for cell in row:
            print('cell',cell)
            row_list.append(cell.value)
        cell_list.append(row_list)
    return cell_list

contents=readExcel('2007.xlsx')
print(contents)


def addSheet(path):

    sheet_name='test Excel4'
    ws=openpyxl.load_workbook(path)
    print(ws.sheetnames)
    if not sheet_name in ws.sheetnames:
        ws.create_sheet(sheet_name,1)
    else:
        print('this sheetname is had')
    ws.save(path)

addSheet('2007.xlsx')